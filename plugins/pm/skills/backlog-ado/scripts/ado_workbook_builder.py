#!/usr/bin/env python3
"""Azure DevOps Workbook Builder.

Generates an Excel workbook (.xlsx) formatted for Azure DevOps CSV import
from a structured JSON backlog data file.

The workbook contains 5 sheets:
    1. Backlog Items - ADO CSV import format with all work item types
    2. Hierarchy Map - Tree visualization (Epic > Feature > Story)
    3. Risk Register - Risk-specific view with assessment fields
    4. Impediments - Impediment items with resolution plans
    5. CI Items - Continuous improvement items

Expected JSON input format:
    {
        "project": "Project Name",
        "area_path": "Project\\Team\\Area",
        "iteration_path": "Project\\Release",
        "work_items": [
            {
                "type": "Epic|Feature|User Story|Technical User Story|Risk|Impediment|CI Item",
                "id": "WI-001",
                "title": "...",
                "description": "...",
                "parent_id": "WI-000 or null",
                "fields": { ... type-specific fields ... }
            }
        ]
    }

Work item field schemas:
    Epic/Feature/User Story/Technical User Story:
        - state: str (New, Active, Resolved, Closed)
        - priority: int (1-4)
        - risk: str (1-High, 2-Medium, 3-Low)
        - story_points: float (stories only)
        - effort: float (epics/features only)
        - business_value: int
        - time_criticality: int
        - value_area: str (Business, Architectural)
        - acceptance_criteria: str (HTML)
        - tags: str (semicolon-separated)
        - start_date: str (YYYY-MM-DD)
        - target_date: str (YYYY-MM-DD)

    Risk:
        - likelihood: int (1-5)
        - impact: str (A-E)
        - risk_assessment: int (1-5)
        - business_impact: str
        - mitigation_plan: str

    Impediment:
        - resolution_plan: str
        - resolved_by: str
        - resolution_date: str

    CI Item:
        - improvement_type: str
        - current_state: str
        - target_state: str
        - action_items: str

Usage:
    python ado_workbook_builder.py --input backlog.json --output backlog.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

BACKLOG_COLUMNS: list[str] = [
    "Work Item Type",
    "ID",
    "Title",
    "State",
    "Area Path",
    "Iteration Path",
    "Description",
    "Acceptance Criteria",
    "Priority",
    "Risk",
    "Story Points",
    "Effort",
    "Business Value",
    "Time Criticality",
    "Parent ID",
    "Value Area",
    "Tags",
    "Start Date",
    "Target Date",
    "Likelihood",
    "Impact",
    "Risk Assessment",
    "Business Impact",
    "Mitigation Plan",
]

RISK_COLUMNS: list[str] = [
    "ID",
    "Title",
    "State",
    "Priority",
    "Likelihood",
    "Impact",
    "Risk Assessment",
    "Business Impact",
    "Mitigation Plan",
    "Description",
    "Tags",
]

IMPEDIMENT_COLUMNS: list[str] = [
    "ID",
    "Title",
    "State",
    "Priority",
    "Description (What is being blocked)",
    "Cause (what is causing the blocker)",
    "Business Impact (Loss of value/Cost of delay)",
    "Possible Solutions",
    "Resolution (What has been done to mitigate)",
    "Area Path",
    "Iteration Path",
    "Tags",
]

CI_COLUMNS: list[str] = [
    "ID",
    "Title",
    "State",
    "Priority",
    "Improvement Type",
    "Current State",
    "Target State",
    "Action Items",
    "Description",
    "Tags",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

WORK_ITEM_TYPES = [
    "Epic",
    "Feature",
    "User Story",
    "Technical User Story",
    "Risk",
    "Impediment",
    "CI Item",
]

TYPE_INDENT: dict[str, int] = {
    "Epic": 0,
    "Feature": 1,
    "User Story": 2,
    "Technical User Story": 2,
    "Risk": 2,
    "Impediment": 2,
    "CI Item": 2,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _load_backlog(input_path: str) -> dict[str, Any]:
    """Load and validate the JSON backlog file."""
    path = Path(input_path)
    if not path.exists():
        print(f"Error: Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    with open(path, encoding="utf-8") as fh:
        data: dict[str, Any] = json.load(fh)

    required_keys = {"project", "work_items"}
    missing = required_keys - data.keys()
    if missing:
        print(f"Error: Missing required keys in JSON: {missing}", file=sys.stderr)
        sys.exit(1)

    return data


def _get_field(item: dict[str, Any], field: str, default: Any = "") -> Any:
    """Retrieve a value from item top-level keys or nested 'fields' dict."""
    if field in item:
        return item[field]
    fields = item.get("fields", {})
    return fields.get(field, default)


def _auto_width(ws: Worksheet) -> None:
    """Set column widths based on the longest cell value per column."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            for cell in row:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_len:
                        max_len = length
        adjusted = min(max_len + 4, 60)
        ws.column_dimensions[col_letter].width = adjusted


def _style_header(ws: Worksheet, columns: list[str]) -> None:
    """Write and style the header row."""
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"


def _style_data_row(ws: Worksheet, row_idx: int, num_cols: int) -> None:
    """Apply alternating row colour and borders to a data row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row_idx % 2 == 0:
            cell.fill = ALT_ROW_FILL


def _add_data_validations(ws: Worksheet, num_rows: int) -> None:
    """Add dropdown data validation for enum columns in the Backlog Items sheet."""
    if num_rows < 2:
        return

    validations: list[tuple[str, str]] = [
        ("I", '"1,2,3,4"'),                          # Priority
        ("J", '"1 - High,2 - Medium,3 - Low"'),      # Risk
        ("P", '"Business,Architectural"'),            # Value Area
        ("D", '"New,Active,Resolved,Closed"'),        # State
    ]
    for col_letter, formula in validations:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please select a valid value."
        dv.errorTitle = "Invalid Entry"
        dv.prompt = "Select from list"
        dv.promptTitle = "Allowed Values"
        cell_range = f"{col_letter}2:{col_letter}{num_rows}"
        dv.add(cell_range)
        ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_backlog_sheet(ws: Worksheet, data: dict[str, Any]) -> None:
    """Populate Sheet 1: Backlog Items (ADO CSV import format)."""
    ws.title = "Backlog Items"
    _style_header(ws, BACKLOG_COLUMNS)

    area_path = data.get("area_path", "")
    iteration_path = data.get("iteration_path", "")

    row_idx = 2
    for item in data["work_items"]:
        wi_type = _get_field(item, "type")
        row_data: list[Any] = [
            wi_type,
            _get_field(item, "id"),
            _get_field(item, "title"),
            _get_field(item, "state", "New"),
            area_path,
            iteration_path,
            _get_field(item, "description"),
            _get_field(item, "acceptance_criteria"),
            _get_field(item, "priority"),
            _get_field(item, "risk"),
            _get_field(item, "story_points"),
            _get_field(item, "effort"),
            _get_field(item, "business_value"),
            _get_field(item, "time_criticality"),
            _get_field(item, "parent_id"),
            _get_field(item, "value_area"),
            _get_field(item, "tags"),
            _get_field(item, "start_date"),
            _get_field(item, "target_date"),
            _get_field(item, "likelihood") if wi_type == "Risk" else "",
            _get_field(item, "impact") if wi_type == "Risk" else "",
            _get_field(item, "risk_assessment") if wi_type == "Risk" else "",
            _get_field(item, "business_impact") if wi_type == "Risk" else "",
            _get_field(item, "mitigation_plan") if wi_type == "Risk" else "",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        _style_data_row(ws, row_idx, len(BACKLOG_COLUMNS))
        row_idx += 1

    _add_data_validations(ws, row_idx - 1)
    _auto_width(ws)


def _build_hierarchy_sheet(ws: Worksheet, data: dict[str, Any]) -> None:
    """Populate Sheet 2: Hierarchy Map (tree view)."""
    ws.title = "Hierarchy Map"
    headers = ["Level", "ID", "Type", "Title", "Parent ID", "Tree Path"]
    _style_header(ws, headers)

    items_by_id: dict[str, dict[str, Any]] = {}
    for item in data["work_items"]:
        items_by_id[_get_field(item, "id")] = item

    row_idx = 2
    for item in data["work_items"]:
        wi_type = _get_field(item, "type")
        indent = TYPE_INDENT.get(wi_type, 0)
        prefix = "  " * indent + ("|- " if indent > 0 else "")
        tree_path = f"{prefix}{_get_field(item, 'title')}"

        row_data = [
            indent,
            _get_field(item, "id"),
            wi_type,
            _get_field(item, "title"),
            _get_field(item, "parent_id"),
            tree_path,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        _style_data_row(ws, row_idx, len(headers))
        row_idx += 1

    _auto_width(ws)


def _build_risk_register_sheet(ws: Worksheet, data: dict[str, Any]) -> None:
    """Populate Sheet 3: Risk Register."""
    ws.title = "Risk Register"
    _style_header(ws, RISK_COLUMNS)

    row_idx = 2
    for item in data["work_items"]:
        if _get_field(item, "type") != "Risk":
            continue
        row_data = [
            _get_field(item, "id"),
            _get_field(item, "title"),
            _get_field(item, "state", "New"),
            _get_field(item, "priority"),
            _get_field(item, "likelihood"),
            _get_field(item, "impact"),
            _get_field(item, "risk_assessment"),
            _get_field(item, "business_impact"),
            _get_field(item, "mitigation_plan"),
            _get_field(item, "description"),
            _get_field(item, "tags"),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        _style_data_row(ws, row_idx, len(RISK_COLUMNS))
        row_idx += 1

    _auto_width(ws)


def _build_impediments_sheet(ws: Worksheet, data: dict[str, Any]) -> None:
    """Populate Sheet 4: Impediments."""
    ws.title = "Impediments"
    _style_header(ws, IMPEDIMENT_COLUMNS)

    row_idx = 2
    for item in data["work_items"]:
        if _get_field(item, "type") != "Impediment":
            continue
        row_data = [
            _get_field(item, "id"),
            _get_field(item, "title"),
            _get_field(item, "state", "New"),
            _get_field(item, "priority"),
            _get_field(item, "description"),
            _get_field(item, "cause"),
            _get_field(item, "business_impact"),
            _get_field(item, "possible_solutions"),
            _get_field(item, "resolution"),
            _get_field(item, "area_path", data.get("area_path", "")),
            _get_field(item, "iteration_path", data.get("iteration_path", "")),
            _get_field(item, "tags"),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        _style_data_row(ws, row_idx, len(IMPEDIMENT_COLUMNS))
        row_idx += 1

    _auto_width(ws)


def _build_ci_sheet(ws: Worksheet, data: dict[str, Any]) -> None:
    """Populate Sheet 5: CI Items."""
    ws.title = "CI Items"
    _style_header(ws, CI_COLUMNS)

    row_idx = 2
    for item in data["work_items"]:
        if _get_field(item, "type") != "CI Item":
            continue
        row_data = [
            _get_field(item, "id"),
            _get_field(item, "title"),
            _get_field(item, "state", "New"),
            _get_field(item, "priority"),
            _get_field(item, "improvement_type"),
            _get_field(item, "current_state"),
            _get_field(item, "target_state"),
            _get_field(item, "action_items"),
            _get_field(item, "description"),
            _get_field(item, "tags"),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        _style_data_row(ws, row_idx, len(CI_COLUMNS))
        row_idx += 1

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def build_workbook(input_path: str, output_path: str) -> None:
    """Build the full ADO workbook from a JSON backlog file.

    Args:
        input_path: Path to the JSON backlog data file.
        output_path: Destination path for the .xlsx file.
    """
    data = _load_backlog(input_path)
    wb = Workbook()

    # Sheet 1 is created by default
    _build_backlog_sheet(wb.active, data)  # type: ignore[arg-type]

    # Sheets 2-5
    _build_hierarchy_sheet(wb.create_sheet(), data)
    _build_risk_register_sheet(wb.create_sheet(), data)
    _build_impediments_sheet(wb.create_sheet(), data)
    _build_ci_sheet(wb.create_sheet(), data)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    print(f"Workbook saved to {output}")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse command-line arguments.

    Args:
        argv: Optional argument list (defaults to sys.argv).

    Returns:
        Parsed namespace with input and output paths.
    """
    parser = argparse.ArgumentParser(
        description="Generate an Azure DevOps CSV-import Excel workbook from JSON backlog data."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to the JSON backlog data file.",
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Path for the output .xlsx file.",
    )
    return parser.parse_args(argv)


if __name__ == "__main__":
    args = parse_args()
    build_workbook(args.input, args.output)
